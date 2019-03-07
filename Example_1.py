#!/usr/bin/env python3

import os

from openpyxl import load_workbook

from Example_tools import menu, check_for_files

def munch(wb, key, munch_column, first_row = 5):
    """Returns a list of tuples sorted by their second element
    There first element of the tuple is a row name
    The second element is the number of hits
    """
    field_value_sum = {}
    for i in wb.sheetnames:
        sheet = wb[i]
        if ((key[0] == None or key[0] == sheet.cell(row=1, column=2).value) and\
            (key[1] == None or key[1] == sheet.cell(row=2, column=2).value)):
            for j in range(first_row, sheet.max_row):
                if(sheet.cell(row=j,column=munch_column).value != None):
                    field = str(sheet.cell(row=j,column=munch_column).value)
                    try:
                        #numbers like 2 in the xlsx file will be read in as 2.0
                        #this will make sure they are read in as 2
                        field = str(int(float(field)))
                    except:
                        pass
                    if field not in field_value_sum:
                        field_value_sum[field] = int(sheet.cell(row=j,column=munch_column+1).value)
                    else:
                        field_value_sum[field] += int(sheet.cell(row=j,column=munch_column+1).value)

    field_value_sum_sorted = sorted(field_value_sum.items(), key=lambda x: x[1], reverse = True)
    return field_value_sum_sorted

if __name__ == '__main__':
    file_list = check_for_files([".xlsx"])
    workbook_filename = file_list[menu(file_list, "Select a file: ")]
    print("--------------------")
    print("Reading from file: " + workbook_filename)
    try:
        wb = load_workbook(workbook_filename)
    except:
        print("Failed to load file, check integrity of file")
        quit()

    try:
        aircraft_list = []
        customer_list = []
        for i in wb.sheetnames:
            sheet = wb[i]
            if str(sheet.cell(row=1,column=2).value) != "Select... ":
                if sheet.cell(row=1,column=2).value not in aircraft_list:
                    aircraft_list.append(int(sheet.cell(row=1,column=2).value))
            if str(sheet.cell(row=2,column=2).value) != "Select...":
                if str(sheet.cell(row=2,column=2).value) not in customer_list:
                    customer_list.append(str(sheet.cell(row=2,column=2).value))
    except:
        print(".xlsx file formating error, check customer and aircraft fields")
        quit()

    key_selected = menu(["Aircrafts","Customers","Both"], "Query by: ")
    if key_selected == 0:
        search_keys = (aircraft_list,None)
    elif key_selected == 1:
        search_keys = (None,customer_list)
    elif key_selected == 2:
        search_keys = (aircraft_list,customer_list)

    key = [None,None]
    if search_keys[0]:
        key[0] = search_keys[0][\
            menu(search_keys[0],"Select a Aircrafts: ")]
    if search_keys[1]:
        key[1] = search_keys[1][\
            menu(search_keys[1],"Select a Customer: ")]

    try:
        format_sheet = wb[wb.sheetnames[0]]
        column_names = ["All"]
        column_name_to_index = {}
        for i in range(1, format_sheet.max_column):
            if format_sheet.cell(row=4, column=i).value == "Hits":
                column_names.append(format_sheet.cell(row=4, column=i-1).value)
                column_name_to_index[format_sheet.cell(row=4, column=i-1).value] = i-1
    except:
        print(".xlsx file formating error, check column headers")
        quit()

    column_to_sum = menu(column_names,"What field do you want to munch: ")

    sum_sorted_fields = {}
    try:
        if column_names[column_to_sum] == "All":
            for i in range(1,len(column_names)-1):
                sum_sorted_fields[column_names[i]] = \
                    munch(wb, key, column_name_to_index[column_names[i]])
        else:
            sum_sorted_fields[column_names[column_to_sum]] = \
                munch(wb, key, column_name_to_index[column_names[column_to_sum]])        
    except:
        print(".xlsx file formating error, check all hit field only have int")
        quit()

    # column_total = {}
    # for i in sumSortedFields:
    #     totalHits = 0
    #     for j in range(0, len(sumSortedFields[i])):
    #         totalHits += sumSortedFields[i][j][1]
    #     fieldTotalHits[i] = totalHits

    # if key[0] and key[1]:
    #     keyPrint = str(key[0]) + " + " + str(key[1])
    # elif key[0]:
    #     keyPrint = str(key[0])
    # elif key[1]:
    #     keyPrint = str(key[1])
    # print("--------------------")
    # print(keyPrint + " traffic looks like:")

    # for i in sumSortedFields:
    #     if sumSortedFields[i]:
    #         print("--------------------")
    #     for j in range(0,len(sumSortedFields[i])):
    #         hitPercentage = ( float(sumSortedFields[i][j][1]) / float(fieldTotalHits[i]) )*100.0
    #         print(i + " " + \
    #                 str(sumSortedFields[i][j][0]) + " had " + \
    #                 str(sumSortedFields[i][j][1]) + " hits ~" + \
    #                 str(round(hitPercentage, 2)) + "% of traffic")

    # print("--------------------")
    # inputKey = raw_input("[1] Save to Output#.Xlsx (Eny other key to exit) ")
    # if inputKey == "1":
    #     n = 1
    #     while os.path.exists("Output%s.xlsx" % n):
    #         n += 1
    #     outputWorkbook = xlsxwriter.Workbook("Output%s.xlsx" % n)
    #     worksheet = outputWorkbook.add_worksheet()
    #     c = 0
    #     for i in sumSortedFields:
    #         if sumSortedFields[i]:
    #             worksheet.write(0, 0, keyPrint)
    #             worksheet.write(1, c+0, i)
    #             worksheet.write(1, c+1, "Hits")
    #             worksheet.write(1, c+2, "%")
    #             for j in range(0,len(sumSortedFields[i])):
    #                 hitPercentage = ( float(sumSortedFields[i][j][1]) / float(fieldTotalHits[i]) )*100.0
    #                 worksheet.write(j+2, c+0, sumSortedFields[i][j][0])
    #                 worksheet.write(j+2, c+1, sumSortedFields[i][j][1])
    #                 worksheet.write(j+2, c+2, round(hitPercentage,2))
    #             c += 3

    #     outputWorkbook.close()
    #     print("Output%s.xlsx Genarated" % n)
